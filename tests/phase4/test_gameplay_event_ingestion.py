from __future__ import annotations

import csv
from dataclasses import replace
from datetime import date
from decimal import Decimal
import json
from pathlib import Path
import tempfile
import unittest

from bi_agent.conversation.store import InMemoryConversationStore
from bi_agent.runtime.contracts import load_contract
from bi_agent.runtime.answer_package import verify_answer_package
from bi_agent.runtime.dataset_catalog import (
    build_dataset_release_authority_record,
    dataset_release_authority_integrity_errors,
    validate_dataset_snapshot_release_payloads,
)
from bi_agent.runtime.runtime_contract_registry import (
    CANONICAL_RUNTIME_BINDINGS_PATH,
    RuntimeContractRegistry,
)
from tools.data.load_gameplay_events_clickhouse import (
    EventLoadError,
    GameplayLoadError,
    build_source_snapshot_payloads,
    load_external_event_workbook,
    load_gameplay_rows,
    load_internal_event_rows,
    persist_source_snapshot_payloads,
    _event_dates,
    _normalize_persisted_row,
    _schema_field_pair,
    _schema_for_part,
    validate_clickhouse_schema,
)
from tools.data.source_loader_common import rows_content_hash


GAMEPLAY_HEADERS = [
    "日期", "区服", "玩法", "游戏人数", "玩法渗透率", "游戏局数", "人均局数",
    "玩家对战率", "总把数", "服务费抽水", "机器人赢的cash", "服务费抽水",
    "系统抽水率", "玩法盈利", "盈利占比", "玩家下注额占比", "玩家下注总额",
    "玩家下注次数", "玩家人均下注次数", "玩家次均下注额", "玩家人均下注额",
    "机器人输的chah",
]

GAMEPLAY_ROW = [
    "2026-06-02", "Waje Special", "Rummy", "10", "0.1", "20", "2", "0.8",
    "20", "12.34", "4", "12.34", "0.001", "5.00", "0.3", "0.4", "5000",
    "50", "5", "100", "500", "2",
]

SHEETS = {
    "尼日利亚宏观通胀汇率波动": "macro_inflation_fx_context",
    "工资周期": "payday_context",
    "重大赛事": "sports_event_context",
    "电力波动": "electricity_issue_context",
    "网络波动": "network_issue_context",
    "投放媒体政策变化": "media_policy_context",
    "极端天气": "weather_context",
    "社会稳定": "social_stability_context",
    "节假日": "holiday_context",
}


class GameplayEventIngestionTest(unittest.TestCase):
    def setUp(self) -> None:
        self.temp = tempfile.TemporaryDirectory()
        self.root = Path(self.temp.name)

    def tearDown(self) -> None:
        self.temp.cleanup()

    def csv_file(self, name: str, rows, *, bom: bool = True) -> Path:
        path = self.root / name
        encoding = "utf-8-sig" if bom else "utf-8"
        with path.open("w", encoding=encoding, newline="") as handle:
            csv.writer(handle).writerows(rows)
        return path

    def gameplay_file(self, name="玩法_2026-01-01_2026-06-02.csv", row=None) -> Path:
        return self.csv_file(name, [GAMEPLAY_HEADERS, row or GAMEPLAY_ROW])

    def normalized_external_workbook(self) -> Path:
        from openpyxl import Workbook

        path = self.root / "events.xlsx"
        workbook = Workbook()
        workbook.remove(workbook.active)
        selected = SHEETS
        for sheet_name in selected:
            business_use = selected[sheet_name]
            sheet = workbook.create_sheet(sheet_name)
            sheet.append(["事件ID", "事件类型", "开始日期", "结束日期", "影响范围", "来源权威", "证据等级", "事件描述"])
            sheet.append([f"{business_use}-1", business_use, "2026-06-01", "2026-06-02", "Nigeria", "reviewed_workbook", "context", "reviewed context"])
        workbook.save(path)
        return path

    def external_workbook(
        self, *, sheets=None, header=None, row=None, salary_rows=None
    ) -> Path:
        from openpyxl import Workbook

        contract = load_contract("contracts/sources/external-events.source.yaml")
        path = self.root / "native-events.xlsx"
        workbook = Workbook()
        workbook.remove(workbook.active)
        selected = sheets or contract["sheet_contracts"]
        for sheet_name in selected:
            sheet = workbook.create_sheet(sheet_name)
            spec = contract["sheet_contracts"].get(sheet_name)
            if spec is None:
                sheet.append(["unexpected"])
                continue
            if spec.get("template_rows_without_header"):
                for salary_row in salary_rows or ("每月 25-30日：主要发薪窗口",):
                    sheet.append([salary_row])
                if row is not None:
                    sheet.append(["每月 25-30日：主要发薪窗口"])
                continue
            for index in range(1, int(spec["header_row"])):
                sheet.append([f"P{index - 1} reviewed legend"])
            headers = list(header or spec["native_headers"])
            sheet.append(headers)
            values = []
            for field_name in headers:
                if field_name == spec["date_column"]:
                    values.append("2026-06-01 至 2026-06-02")
                elif field_name == spec["type_column"]:
                    values.append(spec["business_use"])
                elif field_name == spec["description_column"]:
                    values.append("reviewed context")
                elif field_name == spec.get("scope_column"):
                    values.append("Nigeria")
                else:
                    values.append("reviewed")
            sheet.append(values)
            if row is not None:
                sheet.append(values)
        workbook.save(path)
        return path

    def test_gameplay_loader_uses_utf8_sig_reviewed_duplicate_and_activity_only(self):
        rows, manifest = load_gameplay_rows((self.gameplay_file(),), (), snapshot_id="gameplay-20260602")
        row = rows.overall_rows[0]
        self.assertEqual(row["gameplay"], "Rummy")
        self.assertEqual(row["player_bet_amount"], Decimal("5000.000000000000"))
        self.assertEqual(row["service_fee_rake"], Decimal("12.340000000000"))
        self.assertEqual(row["robot_cash_lost_raw"], Decimal("2.000000000000"))
        self.assertNotIn("paid_amount", row)
        self.assertNotIn("payment_amount", row)
        self.assertEqual(manifest.watermark, "2026-06-02")

    def test_gameplay_duplicate_reviewed_header_rejects_conflicting_semantics(self):
        row = list(GAMEPLAY_ROW)
        row[11] = "99.99"
        with self.assertRaisesRegex(GameplayLoadError, "duplicate_reviewed_column_conflict"):
            load_gameplay_rows((self.gameplay_file(row=row),), (), snapshot_id="s1")

    def test_gameplay_header_filename_shape_numeric_and_scope_fail_closed(self):
        cases = []
        cases.append(("unexpected.csv", [GAMEPLAY_HEADERS, GAMEPLAY_ROW], "source_filename_invalid"))
        cases.append(("玩法_2026-06-03_2026-06-02.csv", [GAMEPLAY_HEADERS, GAMEPLAY_ROW], "source_filename_window_invalid"))
        cases.append(("玩法_2026-01-01_2026-06-02.csv", [GAMEPLAY_HEADERS[:-1], GAMEPLAY_ROW[:-1]], "source_headers_invalid"))
        cases.append(("玩法_2026-01-01_2026-06-02.csv", [GAMEPLAY_HEADERS, GAMEPLAY_ROW[:-1]], "source_row_width"))
        cases.append(("玩法_2026-01-01_2026-06-02.csv", [GAMEPLAY_HEADERS, [*GAMEPLAY_ROW, "extra"]], "source_row_width"))
        invalid_numeric = list(GAMEPLAY_ROW); invalid_numeric[16] = "NaN"
        cases.append(("玩法_2026-01-01_2026-06-02.csv", [GAMEPLAY_HEADERS, invalid_numeric], "source_numeric_invalid"))
        invalid_scope = list(GAMEPLAY_ROW); invalid_scope[1] = "Unknown"
        cases.append(("玩法_2026-01-01_2026-06-02.csv", [GAMEPLAY_HEADERS, invalid_scope], "source_scope_invalid"))
        for index, (name, rows, message) in enumerate(cases):
            with self.subTest(case=message):
                target = name if message.startswith("source_filename") else "玩法_2026-01-01_2026-06-02.csv"
                path = self.csv_file(target, rows)
                with self.assertRaisesRegex(GameplayLoadError, message):
                    load_gameplay_rows((path,), (), snapshot_id="s1")

    def test_gameplay_overall_channel_grains_and_no_data_are_explicit(self):
        overall = self.gameplay_file()
        channel = self.gameplay_file("wajeh5_2026-01-01_2026-06-02.csv")
        empty = self.csv_file("empty_2026-01-01_2026-06-02.csv", [GAMEPLAY_HEADERS])
        rows, manifest = load_gameplay_rows((overall,), (channel, empty), snapshot_id="s1")
        self.assertEqual(tuple(rows.overall_rows[0])[:5], ("snapshot_id", "load_revision", "business_date", "service_scope", "gameplay"))
        self.assertEqual(tuple(rows.channel_rows[0])[:6], ("snapshot_id", "load_revision", "business_date", "channel", "service_scope", "gameplay"))
        self.assertEqual(rows.channel_rows[0]["channel"], "wajeh5")
        self.assertEqual(manifest.no_data_partitions, ("empty",))

    def test_gameplay_duplicate_business_key_uses_contract_aggregation(self):
        first = self.gameplay_file("玩法_2026-01-01_2026-06-02.csv")
        second = self.gameplay_file("玩法_2025-01-01_2025-12-31.csv", ["2025-12-31", *GAMEPLAY_ROW[1:]])
        rows, _ = load_gameplay_rows((first, second), (), snapshot_id="s1")
        self.assertEqual(len(rows.overall_rows), 2)
        duplicate = self.csv_file("玩法_2026-01-01_2026-06-02-copy.csv", [GAMEPLAY_HEADERS, GAMEPLAY_ROW])
        with self.assertRaisesRegex(GameplayLoadError, "source_filename_invalid"):
            load_gameplay_rows((first, duplicate), (), snapshot_id="s1")

    def test_gameplay_duplicate_grain_recomputes_derived_ratios_from_additive_components(self):
        second = list(GAMEPLAY_ROW)
        second[3] = "20"
        second[5] = "60"
        second[6] = "3"
        second[16] = "3000"
        second[12] = "0.001666666666666667"
        second[17] = "30"
        second[18] = "1.5"
        second[19] = "100"
        second[20] = "150"
        source = self.csv_file(
            "玩法_2026-01-01_2026-06-02.csv",
            [GAMEPLAY_HEADERS, GAMEPLAY_ROW, second],
        )
        rows, _ = load_gameplay_rows((source,), (), snapshot_id="s1")
        row = rows.overall_rows[0]
        self.assertEqual(row["gameplay_users"], Decimal("30"))
        self.assertEqual(row["player_bet_amount"], Decimal("8000.000000000000"))
        self.assertEqual(row["rounds_per_user"], Decimal("2.666666666666666667"))
        self.assertEqual(row["player_bet_count_per_user"], Decimal("2.666666666666666667"))
        self.assertEqual(row["player_avg_bet_amount"], Decimal("100.000000000000000000"))

    def test_external_workbook_requires_exact_reviewed_sheets_and_columns(self):
        rows, manifest = load_external_event_workbook(self.external_workbook(), snapshot_id="events-1")
        self.assertEqual(len(rows.event_rows), 9)
        self.assertEqual(
            [(row["source_family"], row["event_id"]) for row in rows.event_rows],
            sorted((row["source_family"], row["event_id"]) for row in rows.event_rows),
        )
        self.assertEqual(rows.event_rows[0]["source_family"], "external_event")
        self.assertEqual(rows.event_rows[0]["wording_limit"], "context")
        payday = next(row for row in rows.event_rows if row["event_type"] == "payday_context")
        self.assertEqual(
            (payday["recurrence_kind"], payday["recurrence_day_start"], payday["recurrence_day_end"]),
            ("monthly_day_range", 25, 30),
        )
        self.assertEqual(json.loads(payday["payload"])["recurrence"]["kind"], "monthly_day_range")
        part = manifest.parts[0]
        normalized = _normalize_persisted_row(
            payday,
            part,
            {field: data_type for field, data_type in map(_schema_field_pair, _schema_for_part(part))},
        )
        self.assertEqual(normalized["recurrence_day_start"], 25)
        self.assertIsInstance(normalized["recurrence_day_start"], int)
        self.assertEqual(manifest.evidence_state, "context_only")
        missing = dict(SHEETS); missing.pop("节假日")
        with self.assertRaisesRegex(EventLoadError, "external_event_sheet_set"):
            load_external_event_workbook(self.external_workbook(sheets=missing), snapshot_id="events-2")
        unexpected = {**SHEETS, "额外表": "other"}
        with self.assertRaisesRegex(EventLoadError, "external_event_sheet_set"):
            load_external_event_workbook(self.external_workbook(sheets=unexpected), snapshot_id="events-3")
        with self.assertRaisesRegex(EventLoadError, "external_event_columns"):
            load_external_event_workbook(self.external_workbook(header=["事件ID", "开始日期"]), snapshot_id="events-4")

    def test_salary_recurrence_is_parsed_per_reviewed_rule(self):
        salary_rules = (
            "每月 23-25日：工资预期窗口",
            "每月 25-30日：主要发薪窗口",
            "次月 1-5日：延迟发薪/补发窗口",
            "每月 10-20日：现金流回落窗口",
            "12月 20日-1月5日：年终工资+节日消费窗口",
        )
        rows, _ = load_external_event_workbook(
            self.external_workbook(salary_rows=salary_rules),
            snapshot_id="events-salary-rules",
        )
        salary = tuple(
            row
            for row in rows.event_rows
            if json.loads(row["payload"])["sheet"] == "工资周期"
        )
        self.assertEqual(len(salary), 5)
        by_description = {
            json.loads(row["payload"])["description"]: (
                row["recurrence_kind"],
                row["recurrence_month_start"],
                row["recurrence_day_start"],
                row["recurrence_month_end"],
                row["recurrence_day_end"],
            )
            for row in salary
        }
        self.assertEqual(
            by_description,
            {
                salary_rules[0]: ("monthly_day_range", 0, 23, 0, 25),
                salary_rules[1]: ("monthly_day_range", 0, 25, 0, 30),
                salary_rules[2]: ("monthly_day_range", 0, 1, 0, 5),
                salary_rules[3]: ("monthly_day_range", 0, 10, 0, 20),
                salary_rules[4]: ("annual_month_day_range", 12, 20, 1, 5),
            },
        )

    def test_external_native_format_rejects_normalized_fallback_and_swapped_sheet_schema(self):
        native = self.external_workbook()
        rows, _ = load_external_event_workbook(native, snapshot_id="native-1")
        self.assertEqual(len(rows.event_rows), 9)
        with self.assertRaisesRegex(EventLoadError, "external_event_columns"):
            load_external_event_workbook(
                self.normalized_external_workbook(), snapshot_id="normalized-rejected"
            )

        from openpyxl import load_workbook

        workbook = load_workbook(native)
        macro = workbook["尼日利亚宏观通胀汇率波动"]
        electricity = workbook["电力波动"]
        for column, cell in enumerate(electricity[1], 1):
            macro.cell(1, column).value = cell.value
        workbook.save(native)
        with self.assertRaisesRegex(EventLoadError, "external_event_columns"):
            load_external_event_workbook(native, snapshot_id="swapped-rejected")

    def test_external_workbook_normalizes_merged_blank_header_and_date_cells(self):
        from datetime import date, datetime
        from openpyxl import load_workbook

        path = self.external_workbook()
        workbook = load_workbook(path)
        sheet = workbook["节假日"]
        sheet["A5"] = "2026-06-01 至 2026-06-02"
        sheet["A1"] = "节假日背景"
        sheet.merge_cells("A1:D1")
        workbook.save(path)
        rows, _ = load_external_event_workbook(path, snapshot_id="events-1")
        holiday = next(row for row in rows.event_rows if row["event_type"] == "holiday_context")
        self.assertEqual(holiday["event_start_date"], "2026-06-01")
        self.assertEqual(holiday["event_end_date"], "2026-06-02")

    def test_external_event_date_windows_cover_reviewed_month_and_quarter_forms(self):
        self.assertEqual(
            _event_dates("2024-08 至 09月", None)[:2],
            (date(2024, 8, 1), date(2024, 9, 30)),
        )
        self.assertEqual(
            _event_dates("2026 Q1整体", None)[:2],
            (date(2026, 1, 1), date(2026, 3, 31)),
        )

    def test_external_event_identity_and_claim_boundary_fail_closed(self):
        path = self.external_workbook(row=["duplicate"])
        with self.assertRaisesRegex(EventLoadError, "event_identity_duplicate"):
            load_external_event_workbook(path, snapshot_id="events-1")
        source = load_contract("contracts/sources/external-events.source.yaml")
        self.assertEqual(source["runtime_boundary"]["evidence_state"], "context_only")
        self.assertEqual(source["runtime_boundary"]["allowed_claim_types"], ["candidate_mechanism"])
        self.assertIn("causal", source["runtime_boundary"]["blocked_claim_types"])
        self.assertIn("roi", source["runtime_boundary"]["blocked_claim_types"])

    def test_internal_event_schema_validates_types_dates_and_stays_unbound(self):
        header = ["event_id", "event_type", "event_start_date", "event_end_date", "scope_type", "scope_value", "authority", "evidence_level", "wording_limit", "description"]
        valid = ["ops-1", "campaign_change", "2026-06-01", "2026-06-02", "global", "global", "operations_owner", "reviewed", "candidate_mechanism", "context"]
        rows, manifest = load_internal_event_rows(self.csv_file("ops.csv", [header, valid]), snapshot_id="ops-1")
        self.assertEqual(rows.event_rows[0]["source_family"], "internal_operation_event")
        self.assertEqual(manifest.evidence_state, "context_only")
        validate_dataset_snapshot_release_payloads(
            build_source_snapshot_payloads(manifest)
        )
        invalids = [
            (["event_id", "event_start_date"], ["e1", "2026-06-02"], "missing_internal_event_fields"),
            (header, [*valid[:2], "2026-06-03", "2026-06-02", *valid[4:]], "internal_event_date_range"),
            (header, [*valid[:7], "strong", *valid[8:]], "internal_event_evidence_level"),
            (header, [*valid[:8], "confirmed_cause", *valid[9:]], "internal_event_wording_limit"),
            (header, [valid[0], "totally_unreviewed_type", *valid[2:]], "internal_event_type"),
            (header, [*valid[:4], "free_text", "anywhere", *valid[6:]], "internal_event_scope"),
        ]
        for index, (fields, values, message) in enumerate(invalids):
            with self.subTest(message=message), self.assertRaisesRegex(EventLoadError, message):
                load_internal_event_rows(self.csv_file(f"bad-{index}.csv", [fields, values]), snapshot_id="ops-x")
        contract = load_contract("contracts/sources/internal-operation-events.source.yaml")
        self.assertEqual(contract["data_contract_state"], "source_unbound")
        self.assertEqual(contract["owner"], "data_operations_owner")
        registry = RuntimeContractRegistry.from_path(CANONICAL_RUNTIME_BINDINGS_PATH)
        self.assertEqual(
            registry.dataset("internal_operation_event")["availability_policy"],
            "released_snapshot_required",
        )

    def test_derived_ratio_contract_recomputes_after_aggregation_and_nulls_invalid_components(self):
        contract = load_contract("contracts/sources/gameplay.source.yaml")
        for field in (
            "rounds_per_user",
            "system_rake_rate",
            "player_bet_count_per_user",
            "player_avg_bet_amount",
            "player_bet_amount_per_user",
        ):
            policy = contract["field_contracts"][field]
            self.assertEqual(policy["duplicate_aggregation"], "recompute_from_components")
            self.assertEqual(
                set(policy["derived_from"]),
                {"numerator_field", "denominator_field", "missing_component_policy", "zero_denominator_policy"},
            )

        missing = list(GAMEPLAY_ROW)
        missing[16] = ""
        missing[19] = ""
        missing[20] = ""
        source = self.csv_file(
            "玩法_2026-01-01_2026-06-02.csv",
            [GAMEPLAY_HEADERS, missing, missing],
        )
        rows, _ = load_gameplay_rows((source,), (), snapshot_id="derived-null")
        self.assertIsNone(rows.overall_rows[0]["player_avg_bet_amount"])
        self.assertIsNone(rows.overall_rows[0]["player_bet_amount_per_user"])

        zero = list(GAMEPLAY_ROW)
        zero[16] = "0"
        zero[17] = "0"
        zero[18] = "0"
        zero[19] = "0"
        zero[20] = "0"
        source = self.csv_file(
            "玩法_2026-01-01_2026-06-02.csv",
            [GAMEPLAY_HEADERS, zero],
        )
        rows, _ = load_gameplay_rows((source,), (), snapshot_id="derived-zero")
        self.assertIsNone(rows.overall_rows[0]["player_avg_bet_amount"])

        betting_subset = list(GAMEPLAY_ROW)
        betting_subset[3] = "3"
        betting_subset[5] = "6"
        betting_subset[6] = "2"
        betting_subset[16] = "1000"
        betting_subset[17] = "2020"
        betting_subset[18] = "1010"
        betting_subset[19] = "0.495049504950495050"
        betting_subset[20] = "500"
        betting_subset[12] = "0.005"
        source = self.csv_file(
            "玩法_2026-01-01_2026-06-02.csv",
            [GAMEPLAY_HEADERS, betting_subset],
        )
        rows, _ = load_gameplay_rows((source,), (), snapshot_id="betting-users")
        self.assertEqual(rows.overall_rows[0]["betting_users_derived"], Decimal("2"))
        self.assertEqual(
            rows.overall_rows[0]["player_bet_count_per_user"],
            Decimal("1010.000000000000000000"),
        )

    def test_manifests_are_versioned_content_addressed_and_release_policy_driven(self):
        gameplay_rows, gameplay = load_gameplay_rows((self.gameplay_file(),), (), snapshot_id="g1")
        gameplay_payloads = build_source_snapshot_payloads(gameplay)
        self.assertEqual({p["dataset_id"] for p in gameplay_payloads}, {"gameplay", "gameplay_channel"})
        self.assertTrue(all("__" in p["physical_table"] for p in gameplay_payloads))
        self.assertTrue(all(p["rows_content_hash"] for p in gameplay_payloads))
        validate_dataset_snapshot_release_payloads(gameplay_payloads)
        external_rows, external = load_external_event_workbook(self.external_workbook(), snapshot_id="e1")
        external_payloads = build_source_snapshot_payloads(external)
        self.assertEqual(len(external_payloads), 1)
        validate_dataset_snapshot_release_payloads(external_payloads)
        authority = build_dataset_release_authority_record(external_payloads)
        self.assertEqual(dataset_release_authority_integrity_errors(authority), ())

    def test_atomic_release_is_idempotent_and_preserves_old_active_revision(self):
        _, first = load_external_event_workbook(self.external_workbook(), snapshot_id="e1")
        store = InMemoryConversationStore()
        first_result = persist_source_snapshot_payloads(store, build_source_snapshot_payloads(first))
        repeat = persist_source_snapshot_payloads(store, build_source_snapshot_payloads(first))
        self.assertEqual(repeat.active_refs, first_result.active_refs)
        from openpyxl import load_workbook
        changed_path = self.external_workbook()
        workbook = load_workbook(changed_path)
        workbook["节假日"]["D5"] = "changed"
        workbook.save(changed_path)
        _, changed = load_external_event_workbook(changed_path, snapshot_id="e1")
        second = persist_source_snapshot_payloads(store, build_source_snapshot_payloads(changed))
        records = store.list_dataset_snapshots("external_event")
        self.assertEqual([item["status"] for item in records], ["superseded", "active"])
        self.assertIn(first_result.active_refs[0], second.superseded_refs)

    def test_runtime_registry_exposes_only_reviewed_activity_and_context_capabilities(self):
        registry = RuntimeContractRegistry.from_path(CANONICAL_RUNTIME_BINDINGS_PATH)
        gameplay = registry.dataset("gameplay")
        external = registry.dataset("external_event")
        self.assertTrue(gameplay["requires_release"])
        self.assertTrue(external["requires_release"])
        self.assertNotIn("paid_amount", registry.metric_sources("player_bet_amount"))
        capability = registry.capability_inputs("gameplay_activity_context")
        self.assertEqual(capability["maximum_claim_strength"], "directional")
        self.assertEqual(capability["query_families"], ["gameplay_activity_probe"])
        self.assertNotIn("paid_amount", capability["required_metrics"])
        event = registry.capability_inputs("event_evidence")
        self.assertEqual(event["maximum_claim_strength"], "candidate_mechanism")
        self.assertNotIn("causal", event["supported_claim_types"])
        self.assertNotIn("roi", event["supported_claim_types"])
        with self.assertRaisesRegex(
            KeyError,
            "unknown_metric_source_adapter:paid_amount:gameplay",
        ):
            registry.metric("paid_amount", dataset_id="gameplay")

        verifier = verify_answer_package(
            draft_claims=(
                {
                    "text": "玩法或外部事件已确认造成付费金额变化并产生 ROI。",
                    "claim_strength": "strong",
                    "evidence_refs": ["event:context"],
                },
            ),
            evidence=(
                {
                    "evidence_ref": "event:context",
                    "evidence_type": "candidate_mechanism",
                    "strength": "context_only",
                    "wording_limit": "context",
                    "typed_payload": {},
                    "limitations": ["causal_and_roi_claims_blocked"],
                },
            ),
            visible_limitations=("causal_and_roi_claims_blocked",),
        )
        self.assertEqual(verifier["status"], "failed")

    def test_clickhouse_schema_validation_uses_canonical_descriptor_not_public_field_names(self):
        _, manifest = load_gameplay_rows((self.gameplay_file(),), (), snapshot_id="g1")

        class Result:
            def __init__(self, rows):
                self.rows = rows

            def named_results(self):
                return iter(self.rows)

        class Client:
            def query(self, sql):
                if "system.columns" in sql:
                    rows = []
                    for part in manifest.parts:
                        descriptor = _schema_for_part(part)
                        for position, item in enumerate(descriptor, 1):
                            if item.startswith(("engine:", "order_by:", "canonicalization:")):
                                continue
                            name, data_type = _schema_field_pair(item)
                            rows.append({"table": part.physical_table, "name": name, "type": data_type, "position": position})
                    return Result(rows)
                return Result([
                    {
                        "name": part.physical_table,
                        "engine": "MergeTree",
                        "sorting_key": next(item.split(":", 1)[1] for item in _schema_for_part(part) if item.startswith("order_by:")),
                    }
                    for part in manifest.parts
                ])

        validate_clickhouse_schema(Client(), manifest)

    def test_persisted_decimal_values_restore_contract_scale_before_full_row_hash(self):
        _, manifest = load_gameplay_rows((self.gameplay_file(),), (), snapshot_id="g1")
        part = manifest.parts[0]
        raw = {
            field: (
                format(value.normalize(), "f")
                if isinstance(value, Decimal)
                else value
            )
            for field, value in part.rows[0].items()
        }
        expected_types = dict(
            _schema_field_pair(item)
            for item in _schema_for_part(part)
            if not item.startswith(("engine:", "order_by:", "canonicalization:"))
        )
        normalized = _normalize_persisted_row(raw, part, expected_types)
        self.assertEqual(rows_content_hash((normalized,)), part.rows_content_hash)


if __name__ == "__main__":
    unittest.main()
