#!/usr/bin/env python3
from __future__ import annotations

# This executable bootstraps the repository root before importing project modules.
# ruff: noqa: E402

import argparse
import calendar
import csv
from dataclasses import asdict, dataclass
from datetime import date, datetime, time, timedelta, timezone
from decimal import Decimal, InvalidOperation, ROUND_HALF_EVEN, localcontext
import json
import math
import os
from pathlib import Path
import re
import sys
from typing import Any, Iterable, Mapping, Sequence


ROOT = Path(__file__).resolve().parents[2]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from bi_agent.conversation.postgres_store import PostgresConversationStore
from bi_agent.runtime.clickhouse_runtime import ClickHouseRuntime
from bi_agent.runtime.contracts import load_contract
from bi_agent.runtime.dataset_catalog import dataset_snapshot_release_ref
from tools.data.load_market_dashboard_clickhouse import (
    DockerClickHouseClient,
    SnapshotPersistenceResult,
    persist_dataset_snapshot_payloads,
)
from tools.data.source_loader_common import (
    canonical_json_bytes,
    content_ref,
    file_sha256,
    insert_json_each_row,
    rows_content_hash,
    schema_fingerprint,
)


GAMEPLAY_CONTRACT_PATH = ROOT / "contracts" / "sources" / "gameplay.source.yaml"
EXTERNAL_CONTRACT_PATH = ROOT / "contracts" / "sources" / "external-events.source.yaml"
INTERNAL_CONTRACT_PATH = (
    ROOT / "contracts" / "sources" / "internal-operation-events.source.yaml"
)
DDL_PATH = ROOT / "tools" / "data" / "clickhouse-analysis-sources.sql"
RUNTIME_BINDING_REF = "contracts/runtime/clickhouse-analysis-bindings.yaml@23"
GAMEPLAY_CONTRACT_REF = "contracts/sources/gameplay.source.yaml@0.1"
EXTERNAL_CONTRACT_REF = "contracts/sources/external-events.source.yaml@0.1"
INTERNAL_CONTRACT_REF = "contracts/sources/internal-operation-events.source.yaml@0.1"

GAMEPLAY_TABLE = "gameplay_daily"
GAMEPLAY_CHANNEL_TABLE = "gameplay_channel_daily"
BUSINESS_EVENTS_TABLE = "business_events"
GAMEPLAY_DATASET = "gameplay"
GAMEPLAY_CHANNEL_DATASET = "gameplay_channel"
EXTERNAL_EVENT_DATASET = "external_event"
INTERNAL_EVENT_DATASET = "internal_operation_event"

GAMEPLAY_OVERALL_RE = re.compile(
    r"^玩法_(?P<start>\d{4}-\d{2}-\d{2})_(?P<end>\d{4}-\d{2}-\d{2})\.csv$"
)
GAMEPLAY_CHANNEL_RE = re.compile(
    r"^(?P<channel>.+)_(?P<start>\d{4}-\d{2}-\d{2})_(?P<end>\d{4}-\d{2}-\d{2})\.csv$"
)
TABLE_RE = re.compile(r"^[A-Za-z_][A-Za-z0-9_]*$")


class GameplayLoadError(ValueError):
    pass


class EventLoadError(ValueError):
    pass


@dataclass(frozen=True)
class DatasetLoadPart:
    dataset_id: str
    physical_table: str
    schema_fields: tuple[str, ...]
    schema_fingerprint: str
    row_count: int
    date_range: tuple[str, ...]
    watermark: str
    rows_content_hash: str
    evidence_state: str
    status: str
    rows: tuple[Mapping[str, Any], ...]


@dataclass(frozen=True)
class SourceLoadManifest:
    manifest_ref: str
    snapshot_ref: str
    snapshot_id: str
    load_revision: str
    release_ref: str
    source_family: str
    contract_ref: str
    runtime_binding_ref: str
    canonicalization_version: str
    source_checksums: Mapping[str, str]
    no_data_partitions: tuple[str, ...]
    no_data_partition_windows: tuple[str, ...]
    evidence_state: str
    parts: tuple[DatasetLoadPart, ...]

    @property
    def watermark(self) -> str:
        return self.parts[0].watermark

    @property
    def row_count(self) -> int:
        return self.parts[0].row_count

    @property
    def physical_table(self) -> str:
        return self.parts[0].physical_table

    def to_dict(self) -> dict[str, Any]:
        value = asdict(self)
        for part in value["parts"]:
            part.pop("rows", None)
        return value


@dataclass(frozen=True)
class GameplayRows:
    snapshot_id: str
    load_revision: str
    overall_rows: tuple[Mapping[str, Any], ...]
    channel_rows: tuple[Mapping[str, Any], ...]


@dataclass(frozen=True)
class EventRows:
    snapshot_id: str
    load_revision: str
    event_rows: tuple[Mapping[str, Any], ...]


def load_gameplay_rows(
    overall_paths: Iterable[str | Path],
    channel_paths: Iterable[str | Path] = (),
    *,
    snapshot_id: str,
    source_contract_path: str | Path = GAMEPLAY_CONTRACT_PATH,
) -> tuple[GameplayRows, SourceLoadManifest]:
    _require_snapshot_id(snapshot_id, GameplayLoadError)
    contract = load_contract(source_contract_path)
    specs = _gameplay_field_specs(contract)
    expected_headers = tuple(
        str(item)
        for item in contract["source_files"]["shared_profile"]["reviewed_headers"]
    )
    overall: list[dict[str, Any]] = []
    channel: list[dict[str, Any]] = []
    no_data: list[str] = []
    no_data_windows: list[str] = []
    source_paths: list[Path] = []
    channel_window_ends: list[date] = []
    seen_names: set[str] = set()

    for raw_path in sorted(
        (Path(item) for item in overall_paths), key=lambda item: item.name
    ):
        start, end, _ = _gameplay_filename(raw_path, channel=False)
        _validate_source_path(raw_path, seen_names, GameplayLoadError)
        parsed = _read_gameplay_csv(
            raw_path, specs, expected_headers, snapshot_id, start, end
        )
        if not parsed:
            raise GameplayLoadError(f"overall_source_has_no_data:{raw_path}")
        overall.extend(parsed)
        source_paths.append(raw_path)
    if not overall:
        raise GameplayLoadError("overall_source_has_no_data")

    for raw_path in sorted(
        (Path(item) for item in channel_paths), key=lambda item: item.name
    ):
        start, end, channel_name = _gameplay_filename(raw_path, channel=True)
        _validate_source_path(raw_path, seen_names, GameplayLoadError)
        parsed = _read_gameplay_csv(
            raw_path,
            specs,
            expected_headers,
            snapshot_id,
            start,
            end,
            channel=channel_name,
        )
        channel_window_ends.append(end)
        if parsed:
            channel.extend(parsed)
        else:
            no_data.append(channel_name)
            no_data_windows.append(
                f"{channel_name}@{start.isoformat()}:{end.isoformat()}"
            )
        source_paths.append(raw_path)

    overall = _aggregate_gameplay(
        overall,
        ("snapshot_id", "business_date", "service_scope", "gameplay"),
        specs,
        GAMEPLAY_DATASET,
    )
    channel = _aggregate_gameplay(
        channel,
        ("snapshot_id", "business_date", "channel", "service_scope", "gameplay"),
        specs,
        GAMEPLAY_CHANNEL_DATASET,
    )
    overall_schema = _gameplay_schema(specs, channel=False)
    channel_schema = _gameplay_schema(specs, channel=True)
    overall_fingerprint = schema_fingerprint(overall_schema)
    channel_fingerprint = schema_fingerprint(channel_schema)
    checksums = {
        item.name: file_sha256(item)
        for item in sorted(source_paths, key=lambda item: item.name)
    }
    revision = content_ref(
        "gameplay-load",
        {
            "snapshot_id": snapshot_id,
            "source_checksums": checksums,
            "overall_rows_hash": rows_content_hash(overall),
            "channel_rows_hash": rows_content_hash(channel),
            "overall_schema_fingerprint": overall_fingerprint,
            "channel_schema_fingerprint": channel_fingerprint,
            "canonicalization_version": contract["runtime_binding"][
                "canonicalization_version"
            ],
        },
    )
    overall = _attach_revision(overall, revision)
    channel = _attach_revision(channel, revision)
    overall_range = _row_date_range(overall, "business_date")
    channel_range = _row_date_range(channel, "business_date") if channel else ()
    channel_watermark = (
        channel_range[-1]
        if channel_range
        else max(channel_window_ends).isoformat()
        if channel_window_ends
        else overall_range[-1]
    )
    parts = (
        _part(
            GAMEPLAY_DATASET,
            GAMEPLAY_TABLE,
            overall_schema,
            overall_fingerprint,
            overall,
            overall_range,
            overall_range[-1],
            evidence_state="context_only",
        ),
        _part(
            GAMEPLAY_CHANNEL_DATASET,
            GAMEPLAY_CHANNEL_TABLE,
            channel_schema,
            channel_fingerprint,
            channel,
            channel_range,
            channel_watermark,
            evidence_state="context_only",
            status="active" if channel else "no_data",
        ),
    )
    manifest = _build_manifest(
        snapshot_id=snapshot_id,
        load_revision=revision,
        source_family="gameplay_activity",
        contract_ref=GAMEPLAY_CONTRACT_REF,
        canonicalization_version=contract["runtime_binding"][
            "canonicalization_version"
        ],
        source_checksums=checksums,
        no_data_partitions=tuple(sorted(set(no_data))),
        no_data_partition_windows=tuple(sorted(no_data_windows)),
        evidence_state="context_only",
        parts=parts,
    )
    return GameplayRows(snapshot_id, revision, tuple(overall), tuple(channel)), manifest


def load_external_event_workbook(
    workbook_path: str | Path,
    *,
    snapshot_id: str,
    source_contract_path: str | Path = EXTERNAL_CONTRACT_PATH,
) -> tuple[EventRows, SourceLoadManifest]:
    _require_snapshot_id(snapshot_id, EventLoadError)
    path = Path(workbook_path)
    if not path.is_file():
        raise EventLoadError(f"source_file_missing:{path}")
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise EventLoadError("openpyxl_dependency_missing") from exc
    contract = load_contract(source_contract_path)
    sheet_contracts = contract.get("sheet_contracts")
    if not isinstance(sheet_contracts, Mapping) or len(sheet_contracts) != 9:
        raise EventLoadError("external_event_sheet_contracts")
    workbook = load_workbook(path, read_only=False, data_only=True)
    if (
        set(workbook.sheetnames) != set(sheet_contracts)
        or len(workbook.sheetnames) != 9
    ):
        raise EventLoadError("external_event_sheet_set")
    rows: list[dict[str, Any]] = []
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        sheet_spec = sheet_contracts[sheet_name]
        rows.extend(
            _read_external_sheet(sheet, sheet_name, sheet_spec, contract, snapshot_id)
        )
    if not rows:
        raise EventLoadError("external_event_workbook_no_data")
    _validate_event_identities(rows)
    rows.sort(key=lambda item: (str(item["source_family"]), str(item["event_id"])))
    revision = content_ref(
        "external-events-load",
        {
            "snapshot_id": snapshot_id,
            "source_checksum": file_sha256(path),
            "rows_hash": rows_content_hash(rows),
            "canonicalization_version": contract["runtime_binding"][
                "canonicalization_version"
            ],
        },
    )
    rows = _attach_revision(rows, revision)
    schema = _event_schema(contract["runtime_binding"]["canonicalization_version"])
    fingerprint = schema_fingerprint(schema)
    coverage_end = str(contract["coverage"]["date_range"]["end"])
    parts = (
        _part(
            EXTERNAL_EVENT_DATASET,
            BUSINESS_EVENTS_TABLE,
            schema,
            fingerprint,
            rows,
            _row_date_range(rows, "event_start_date"),
            coverage_end,
            evidence_state="context_only",
        ),
    )
    manifest = _build_manifest(
        snapshot_id=snapshot_id,
        load_revision=revision,
        source_family="external_event",
        contract_ref=EXTERNAL_CONTRACT_REF,
        canonicalization_version=contract["runtime_binding"][
            "canonicalization_version"
        ],
        source_checksums={path.name: file_sha256(path)},
        no_data_partitions=(),
        no_data_partition_windows=(),
        evidence_state="context_only",
        parts=parts,
    )
    return EventRows(snapshot_id, revision, tuple(rows)), manifest


def load_internal_event_rows(
    csv_path: str | Path,
    *,
    snapshot_id: str,
    source_contract_path: str | Path = INTERNAL_CONTRACT_PATH,
) -> tuple[EventRows, SourceLoadManifest]:
    _require_snapshot_id(snapshot_id, EventLoadError)
    path = Path(csv_path)
    if not path.is_file():
        raise EventLoadError(f"source_file_missing:{path}")
    contract = load_contract(source_contract_path)
    schema_contract = contract["input_schema"]
    with path.open(encoding="utf-8-sig", newline="") as handle:
        reader = csv.reader(handle)
        try:
            headers = next(reader)
        except StopIteration as exc:
            raise EventLoadError("missing_internal_event_fields") from exc
        required = tuple(schema_contract["required_fields"])
        optional = tuple(schema_contract["optional_fields"])
        missing = tuple(field for field in required if field not in headers)
        if missing:
            raise EventLoadError("missing_internal_event_fields:" + ",".join(missing))
        if len(headers) != len(set(headers)) or set(headers) - set(
            (*required, *optional)
        ):
            raise EventLoadError("internal_event_columns")
        rows = []
        for row_number, values in enumerate(reader, 2):
            if len(values) != len(headers):
                raise EventLoadError(f"source_row_width:{path}:{row_number}")
            if not any(str(item).strip() for item in values):
                continue
            raw = dict(zip(headers, values))
            start = _strict_date(raw["event_start_date"], "event_start_date")
            end = _strict_date(raw["event_end_date"], "event_end_date")
            if start > end:
                raise EventLoadError("internal_event_date_range")
            for field in required:
                if not str(raw.get(field) or "").strip():
                    raise EventLoadError(f"missing_internal_event_fields:{field}")
            _require_allowed(
                raw["event_type"],
                schema_contract["event_type_values"],
                "internal_event_type",
            )
            scope_model = schema_contract.get("affected_scope_model")
            scope_type = str(raw["scope_type"])
            scope_value = str(raw["scope_value"])
            if (
                not isinstance(scope_model, Mapping)
                or scope_type not in scope_model
                or scope_value
                not in tuple(scope_model[scope_type].get("allowed_values") or ())
            ):
                raise EventLoadError("internal_event_scope")
            _require_allowed(
                raw["authority"],
                schema_contract["authority_values"],
                "internal_event_authority",
            )
            _require_allowed(
                raw["evidence_level"],
                schema_contract["evidence_level_values"],
                "internal_event_evidence_level",
            )
            _require_allowed(
                raw["wording_limit"],
                schema_contract["wording_limit_values"],
                "internal_event_wording_limit",
            )
            rows.append(
                _event_row(
                    snapshot_id=snapshot_id,
                    source_family="internal_operation_event",
                    event_id=raw["event_id"],
                    event_type=raw["event_type"],
                    start=start,
                    end=end,
                    scope=f"{scope_type}:{scope_value}",
                    authority=raw["authority"],
                    evidence_level=raw["evidence_level"],
                    wording_limit=raw["wording_limit"],
                    payload={"description": raw.get("description", "")},
                )
            )
    if not rows:
        raise EventLoadError("internal_event_source_no_data")
    _validate_event_identities(rows)
    rows.sort(key=lambda item: (str(item["source_family"]), str(item["event_id"])))
    revision = content_ref(
        "internal-events-load",
        {
            "snapshot_id": snapshot_id,
            "source_checksum": file_sha256(path),
            "rows_hash": rows_content_hash(rows),
        },
    )
    rows = _attach_revision(rows, revision)
    schema = _event_schema(contract["runtime_binding"]["canonicalization_version"])
    fingerprint = schema_fingerprint(schema)
    event_range = _row_date_range(rows, "event_start_date")
    parts = (
        _part(
            INTERNAL_EVENT_DATASET,
            BUSINESS_EVENTS_TABLE,
            schema,
            fingerprint,
            rows,
            event_range,
            event_range[-1],
            evidence_state="context_only",
        ),
    )
    manifest = _build_manifest(
        snapshot_id=snapshot_id,
        load_revision=revision,
        source_family="internal_operation_event",
        contract_ref=INTERNAL_CONTRACT_REF,
        canonicalization_version=contract["runtime_binding"][
            "canonicalization_version"
        ],
        source_checksums={path.name: file_sha256(path)},
        no_data_partitions=(),
        no_data_partition_windows=(),
        evidence_state="context_only",
        parts=parts,
    )
    return EventRows(snapshot_id, revision, tuple(rows)), manifest


def build_source_snapshot_payloads(
    manifest: SourceLoadManifest,
) -> tuple[dict[str, Any], ...]:
    payloads = []
    for part in manifest.parts:
        payloads.append(
            {
                "snapshot_ref": _snapshot_ref(manifest, part),
                "dataset_id": part.dataset_id,
                "physical_table": part.physical_table,
                "watermark": part.watermark,
                "schema_fingerprint": part.schema_fingerprint,
                "schema_fields": list(part.schema_fields),
                "contract_ref": manifest.contract_ref,
                "loaded_at": _snapshot_available_at(part.watermark),
                "status": part.status,
                "snapshot_id": manifest.snapshot_id,
                "logical_snapshot_id": manifest.snapshot_id,
                "load_revision": manifest.load_revision,
                "release_ref": manifest.release_ref,
                "requires_release": True,
                "evidence_state": part.evidence_state,
                "reconciliation_status": "not_applicable",
                "reconciliation_ref": content_ref(
                    "source-reconciliation",
                    {"dataset_id": part.dataset_id, "status": "not_applicable"},
                ),
                "source_load_manifest_ref": manifest.manifest_ref,
                "runtime_binding_ref": manifest.runtime_binding_ref,
                "source_checksums": dict(manifest.source_checksums),
                "no_data_partitions": list(manifest.no_data_partitions),
                "no_data_partition_windows": list(manifest.no_data_partition_windows),
                "row_count": part.row_count,
                "date_range": list(part.date_range),
                "rows_content_hash": part.rows_content_hash,
            }
        )
    return tuple(payloads)


def persist_source_snapshot_payloads(
    store: Any, payloads: Sequence[Mapping[str, Any]]
) -> SnapshotPersistenceResult:
    try:
        return persist_dataset_snapshot_payloads(store, payloads)
    except ValueError as exc:
        raise EventLoadError(f"postgres_release_preflight:{exc}") from exc


def stage_source_release(
    client: Any,
    manifest: SourceLoadManifest,
    *,
    active_load_revisions: Sequence[str],
) -> str:
    staged = False
    for part in manifest.parts:
        existing = _read_persisted(
            client, part, manifest.snapshot_id, manifest.load_revision
        )
        if existing:
            try:
                _validate_persisted_part(existing, part)
                continue
            except ValueError as exc:
                if manifest.load_revision in set(active_load_revisions):
                    raise GameplayLoadError("active_load_revision_invalid") from exc
                client.command(
                    f"DELETE FROM {part.physical_table} WHERE snapshot_id = {{snapshot_id:String}} AND load_revision = {{load_revision:String}}",
                    parameters={
                        "snapshot_id": manifest.snapshot_id,
                        "load_revision": manifest.load_revision,
                    },
                    settings={"mutations_sync": 2},
                )
        for offset in range(0, len(part.rows), 10_000):
            insert_json_each_row(
                client,
                part.physical_table,
                part.rows[offset : offset + 10_000],
            )
        _validate_persisted_part(
            _read_persisted(client, part, manifest.snapshot_id, manifest.load_revision),
            part,
        )
        staged = True
    return "staged_and_validated" if staged else "already_validated"


def apply_clickhouse_ddl(client: Any, manifest: SourceLoadManifest) -> None:
    ddl = (
        DDL_PATH.read_text(encoding="utf-8")
        .split("-- BEGIN GAMEPLAY_EVENTS", 1)[1]
        .split("-- END GAMEPLAY_EVENTS", 1)[0]
    )
    replacements = {
        "__GAMEPLAY_TABLE__": next(
            (
                part.physical_table
                for part in manifest.parts
                if part.dataset_id == GAMEPLAY_DATASET
            ),
            _versioned_table(GAMEPLAY_TABLE, _gameplay_current_fingerprint(False)),
        ),
        "__GAMEPLAY_CHANNEL_TABLE__": next(
            (
                part.physical_table
                for part in manifest.parts
                if part.dataset_id == GAMEPLAY_CHANNEL_DATASET
            ),
            _versioned_table(
                GAMEPLAY_CHANNEL_TABLE, _gameplay_current_fingerprint(True)
            ),
        ),
        "__BUSINESS_EVENTS_TABLE__": next(
            (
                part.physical_table
                for part in manifest.parts
                if part.dataset_id in {EXTERNAL_EVENT_DATASET, INTERNAL_EVENT_DATASET}
            ),
            _versioned_table(BUSINESS_EVENTS_TABLE, _event_current_fingerprint()),
        ),
    }
    for token, table in replacements.items():
        if not TABLE_RE.fullmatch(table):
            raise GameplayLoadError("invalid_clickhouse_table")
        ddl = ddl.replace(token, table)
    required_tables = {part.physical_table for part in manifest.parts}
    for statement in (item.strip() for item in ddl.split(";")):
        if statement and any(
            f"TABLE IF NOT EXISTS {table}" in statement for table in required_tables
        ):
            client.command(statement)
    validate_clickhouse_schema(client, manifest)


def validate_clickhouse_schema(client: Any, manifest: SourceLoadManifest) -> None:
    tables = tuple(part.physical_table for part in manifest.parts)
    result = client.query(
        "SELECT table, name, type, position FROM system.columns WHERE database = currentDatabase() AND table IN ("
        + ",".join(f"'{table}'" for table in tables)
        + ") ORDER BY table, position"
    )
    observed_rows = tuple(dict(item) for item in result.named_results())
    for part in manifest.parts:
        observed = tuple(
            (str(item["name"]), str(item["type"]))
            for item in observed_rows
            if item.get("table") == part.physical_table
        )
        expected = tuple(
            _schema_field_pair(item)
            for item in _schema_for_part(part)
            if not item.startswith(("engine:", "order_by:", "canonicalization:"))
        )
        if observed != expected:
            raise GameplayLoadError(
                f"clickhouse_schema_drift:columns:{part.dataset_id}"
            )
    table_result = client.query(
        "SELECT name, engine, sorting_key FROM system.tables WHERE database = currentDatabase() AND name IN ("
        + ",".join(f"'{table}'" for table in tables)
        + ") ORDER BY name"
    )
    by_name = {str(item["name"]): dict(item) for item in table_result.named_results()}
    for part in manifest.parts:
        order = next(
            item.split(":", 1)[1]
            for item in _schema_for_part(part)
            if item.startswith("order_by:")
        )
        item = by_name.get(part.physical_table, {})
        sorting = re.sub(r"[()`\s]", "", str(item.get("sorting_key") or ""))
        if item.get("engine") != "MergeTree" or sorting != order:
            raise GameplayLoadError(f"clickhouse_schema_drift:table:{part.dataset_id}")


def _gameplay_field_specs(contract: Mapping[str, Any]) -> dict[str, Mapping[str, Any]]:
    specs = contract.get("field_contracts")
    source_fields = set(contract.get("field_mapping", {})) - {
        "service_fee_rake_raw_columns",
        "service_fee_rake_selected_column",
        "service_fee_rake_selection_reason",
        "ggr",
    } | {"service_fee_rake"}
    if (
        not isinstance(specs, Mapping)
        or not source_fields.issubset(specs)
        or any(
            field not in source_fields
            and (
                not isinstance(spec, Mapping)
                or spec.get("source_field") is not None
                or not isinstance(spec.get("derivation"), Mapping)
            )
            for field, spec in specs.items()
        )
    ):
        raise GameplayLoadError("gameplay_field_contracts_invalid")
    if any(not isinstance(value, Mapping) for value in specs.values()):
        raise GameplayLoadError("gameplay_field_contracts_invalid")
    return {str(key): dict(value) for key, value in specs.items()}


def _read_gameplay_csv(
    path: Path,
    specs: Mapping[str, Mapping[str, Any]],
    expected_headers: tuple[str, ...],
    snapshot_id: str,
    start: date,
    end: date,
    *,
    channel: str = "",
) -> list[dict[str, Any]]:
    with path.open(encoding="utf-8-sig", newline="") as handle:
        reader = csv.reader(handle)
        try:
            headers = next(reader)
        except StopIteration:
            raise GameplayLoadError(f"source_headers_invalid:{path}")
        service_positions = [
            index for index, value in enumerate(headers) if value == "服务费抽水"
        ]
        duplicate_occurrences = tuple(
            int(item)
            for item in specs["service_fee_rake"]["reviewed_duplicate_occurrences"]
        )
        if tuple(headers) != expected_headers or len(service_positions) != len(
            duplicate_occurrences
        ):
            raise GameplayLoadError(f"source_headers_invalid:{path}")
        positions: dict[str, int] = {}
        for field, spec in specs.items():
            if spec.get("source_field") is None:
                continue
            source_name = str(spec["source_field"])
            occurrence = int(spec.get("source_occurrence", 1))
            candidates = [
                index for index, value in enumerate(headers) if value == source_name
            ]
            if len(candidates) < occurrence:
                raise GameplayLoadError(f"source_headers_invalid:{path}")
            positions[field] = candidates[occurrence - 1]
        rows = []
        for row_number, values in enumerate(reader, 2):
            if len(values) != len(headers):
                raise GameplayLoadError(f"source_row_width:{path}:{row_number}")
            if not any(value.strip() for value in values):
                continue
            parsed_date = _strict_date(
                values[positions["business_date"]], "business_date", GameplayLoadError
            )
            if not start <= parsed_date <= end:
                raise GameplayLoadError(
                    f"source_date_outside_filename_window:{path}:{row_number}"
                )
            first_rake = _parse_decimal(
                values[service_positions[0]],
                specs["service_fee_rake"],
                path,
                row_number,
            )
            second_rake = _parse_decimal(
                values[service_positions[1]],
                specs["service_fee_rake"],
                path,
                row_number,
            )
            if first_rake != second_rake:
                raise GameplayLoadError(
                    f"duplicate_reviewed_column_conflict:{path}:{row_number}"
                )
            row: dict[str, Any] = {
                "snapshot_id": snapshot_id,
                "business_date": parsed_date.isoformat(),
            }
            if channel:
                row["channel"] = channel
            for field, spec in specs.items():
                if field == "business_date" or spec.get("source_field") is None:
                    continue
                value = values[positions[field]]
                if spec["logical_type"] == "string":
                    normalized = value.strip()
                    if not normalized:
                        raise GameplayLoadError(
                            f"source_string_missing:{field}:{path}:{row_number}"
                        )
                    allowed = spec.get("allowed_values")
                    if allowed and normalized not in allowed:
                        raise GameplayLoadError(
                            f"source_scope_invalid:{field}:{normalized}"
                        )
                    row[field] = normalized
                else:
                    row[field] = (
                        first_rake
                        if field == "service_fee_rake"
                        else _parse_decimal(value, spec, path, row_number)
                    )
            row["betting_users_derived"] = _derive_betting_users(
                row,
                specs["betting_users_derived"],
                path,
                row_number,
            )
            _validate_source_derived_values(row, specs, path, row_number)
            rows.append(row)
    return rows


def _parse_decimal(
    value: Any, spec: Mapping[str, Any], path: Path, row_number: int
) -> Decimal | None:
    text = str(value).strip()
    if text in {str(item) for item in spec.get("missing_tokens", ())}:
        if spec.get("nullable") is True:
            return None
        raise GameplayLoadError(f"source_numeric_missing:{path}:{row_number}")
    try:
        number = Decimal(text)
    except (InvalidOperation, ValueError) as exc:
        raise GameplayLoadError(f"source_numeric_invalid:{path}:{row_number}") from exc
    if not number.is_finite():
        raise GameplayLoadError(f"source_numeric_invalid:{path}:{row_number}")
    precision = int(spec["precision"])
    scale = int(spec["scale"])
    with localcontext() as context:
        context.prec = precision + scale + 8
        try:
            number = number.quantize(
                Decimal(1).scaleb(-scale), rounding=ROUND_HALF_EVEN
            )
        except InvalidOperation as exc:
            raise GameplayLoadError(
                f"source_numeric_precision:{path}:{row_number}"
            ) from exc
    digits = len(number.as_tuple().digits)
    integer_digits = max(0, digits + number.as_tuple().exponent)
    if integer_digits > precision - scale:
        raise GameplayLoadError(f"source_numeric_precision:{path}:{row_number}")
    return number


def _aggregate_gameplay(
    rows: Sequence[Mapping[str, Any]],
    key_fields: tuple[str, ...],
    specs: Mapping[str, Mapping[str, Any]],
    dataset_id: str,
) -> list[dict[str, Any]]:
    grouped: dict[tuple[Any, ...], list[Mapping[str, Any]]] = {}
    for item in rows:
        key = tuple(item[field] for field in key_fields)
        grouped.setdefault(key, []).append(item)
    aggregated = []
    for key in sorted(grouped, key=lambda item: tuple(str(value) for value in item)):
        members = grouped[key]
        target = {field: members[0][field] for field in key_fields}
        for field, spec in specs.items():
            if field in key_fields:
                continue
            policy = spec["duplicate_aggregation"]
            if policy == "additive_sum":
                values = [member.get(field) for member in members]
                target[field] = (
                    None
                    if any(value is None for value in values)
                    else sum((Decimal(value) for value in values), Decimal(0))
                )
            elif policy == "recompute_from_components":
                continue
            elif policy == "weighted_average":
                weighted = [
                    (member.get(field), member.get(spec["weight_field"]))
                    for member in members
                ]
                if any(value is None or weight is None for value, weight in weighted):
                    target[field] = None
                else:
                    total_weight = sum(
                        (Decimal(weight) for _, weight in weighted), Decimal(0)
                    )
                    target[field] = (
                        _quantize_contract_value(
                            sum(
                                (
                                    Decimal(value) * Decimal(weight)
                                    for value, weight in weighted
                                ),
                                Decimal(0),
                            )
                            / total_weight,
                            spec,
                        )
                        if total_weight
                        else None
                    )
            else:
                raise GameplayLoadError(
                    f"duplicate_policy_unsupported:{field}:{policy}"
                )
        for field, spec in specs.items():
            if spec.get("duplicate_aggregation") != "recompute_from_components":
                continue
            derived = spec.get("derived_from")
            if not isinstance(derived, Mapping):
                raise GameplayLoadError(f"derived_ratio_contract_invalid:{field}")
            numerator = target.get(str(derived.get("numerator_field") or ""))
            denominator = target.get(str(derived.get("denominator_field") or ""))
            if numerator is None or denominator is None:
                if derived.get("missing_component_policy") != "null":
                    raise GameplayLoadError(f"derived_ratio_missing_component:{field}")
                target[field] = None
                continue
            denominator_decimal = Decimal(denominator)
            if denominator_decimal <= 0:
                if derived.get("zero_denominator_policy") != "null":
                    raise GameplayLoadError(
                        f"derived_ratio_invalid_denominator:{field}"
                    )
                target[field] = None
                continue
            target[field] = _quantize_contract_value(
                Decimal(numerator) / denominator_decimal,
                spec,
            )
        aggregated.append(target)
    return aggregated


def _validate_source_derived_values(
    row: Mapping[str, Any],
    specs: Mapping[str, Mapping[str, Any]],
    path: Path,
    row_number: int,
) -> None:
    for field, spec in specs.items():
        if spec.get("duplicate_aggregation") != "recompute_from_components":
            continue
        derived = spec.get("derived_from")
        validation = spec.get("source_value_validation")
        if not isinstance(derived, Mapping) or not isinstance(validation, Mapping):
            raise GameplayLoadError(f"derived_ratio_contract_invalid:{field}")
        if validation.get("policy") != "compare_when_components_valid":
            raise GameplayLoadError(f"derived_ratio_validation_policy:{field}")
        numerator = row.get(str(derived.get("numerator_field") or ""))
        denominator = row.get(str(derived.get("denominator_field") or ""))
        source_value = row.get(field)
        if numerator is None or denominator is None or source_value is None:
            continue
        denominator_decimal = Decimal(denominator)
        if denominator_decimal <= 0:
            continue
        expected = _quantize_contract_value(
            Decimal(numerator) / denominator_decimal,
            spec,
        )
        absolute_tolerance = Decimal(str(validation.get("absolute_tolerance") or "0"))
        relative_tolerance = Decimal(str(validation.get("relative_tolerance") or "0"))
        tolerance = max(
            absolute_tolerance,
            abs(expected) * relative_tolerance,
        )
        if abs(Decimal(source_value) - expected) > tolerance:
            raise GameplayLoadError(
                f"derived_ratio_source_mismatch:{field}:{path}:{row_number}"
            )


def _derive_betting_users(
    row: Mapping[str, Any],
    spec: Mapping[str, Any],
    path: Path,
    row_number: int,
) -> Decimal | None:
    derivation = spec.get("derivation")
    if (
        not isinstance(derivation, Mapping)
        or derivation.get("policy") != "reconcile_ratio_denominators"
    ):
        raise GameplayLoadError("betting_users_derivation_contract_invalid")
    tolerance = Decimal(str(derivation.get("tolerance") or "0"))
    candidates: list[Decimal] = []
    for pair in derivation.get("source_pairs") or ():
        if not isinstance(pair, Mapping):
            raise GameplayLoadError("betting_users_derivation_contract_invalid")
        numerator = row.get(str(pair.get("numerator_field") or ""))
        per_user = row.get(str(pair.get("per_user_field") or ""))
        if numerator is None or per_user is None:
            continue
        numerator_decimal = Decimal(numerator)
        per_user_decimal = Decimal(per_user)
        if per_user_decimal == 0:
            if numerator_decimal != 0:
                raise GameplayLoadError(
                    f"betting_users_zero_ratio_conflict:{path}:{row_number}"
                )
            candidates.append(Decimal(0))
            continue
        candidate = numerator_decimal / per_user_decimal
        integral = candidate.quantize(Decimal(1), rounding=ROUND_HALF_EVEN)
        if candidate < 0 or abs(candidate - integral) > tolerance:
            raise GameplayLoadError(f"betting_users_non_integral:{path}:{row_number}")
        candidates.append(integral)
    if not candidates:
        return None
    if any(candidate != candidates[0] for candidate in candidates[1:]):
        raise GameplayLoadError(f"betting_users_source_mismatch:{path}:{row_number}")
    upper_bound = row.get(str(derivation.get("upper_bound_field") or ""))
    if upper_bound is not None and candidates[0] > Decimal(upper_bound):
        raise GameplayLoadError(
            f"betting_users_exceeds_gameplay_users:{path}:{row_number}"
        )
    return candidates[0]


def _quantize_contract_value(value: Decimal, spec: Mapping[str, Any]) -> Decimal:
    scale = int(spec["scale"])
    with localcontext() as context:
        context.prec = int(spec["precision"]) + scale + 8
        return value.quantize(Decimal(1).scaleb(-scale), rounding=ROUND_HALF_EVEN)


def _read_external_sheet(
    sheet: Any,
    sheet_name: str,
    spec: Mapping[str, Any],
    contract: Mapping[str, Any],
    snapshot_id: str,
) -> list[dict[str, Any]]:
    values = [tuple(cell.value for cell in row) for row in sheet.iter_rows()]
    nonblank_rows = [
        (index + 1, row)
        for index, row in enumerate(values)
        if any(value not in (None, "") for value in row)
    ]
    if not nonblank_rows:
        return []
    if spec.get("template_rows_without_header") is True:
        rows = []
        for row_number, row in nonblank_rows:
            description = str(
                row[int(spec["description_column_index"]) - 1] or ""
            ).strip()
            if not description:
                continue
            rows.append(
                _native_event_row(
                    snapshot_id,
                    sheet_name,
                    spec,
                    row_number,
                    date.fromisoformat(spec["recurring_window_start"]),
                    date.fromisoformat(spec["recurring_window_end"]),
                    spec["business_use"],
                    description,
                    spec["scope_default"],
                    row,
                )
            )
        return rows
    header_row = int(spec["header_row"])
    if header_row > len(values):
        raise EventLoadError(f"external_event_columns:{sheet_name}")
    header = tuple(
        str(value).strip() if value is not None else ""
        for value in values[header_row - 1]
    )
    expected = tuple(str(item) for item in spec["native_headers"])
    if header != expected:
        raise EventLoadError(f"external_event_columns:{sheet_name}")
    positions = {name: index for index, name in enumerate(expected)}
    rows = []
    inferred_year: int | None = None
    for offset, raw in enumerate(values[header_row:], header_row + 1):
        if not any(value not in (None, "") for value in raw):
            continue
        start, end, inferred_year = _event_dates(
            raw[positions[spec["date_column"]]], inferred_year
        )
        event_type = str(
            raw[positions[spec["type_column"]]] or spec["business_use"]
        ).strip()
        description = str(raw[positions[spec["description_column"]]] or "").strip()
        scope = (
            str(raw[positions[spec["scope_column"]]] or "").strip()
            if spec.get("scope_column")
            else str(spec["scope_default"])
        )
        if not event_type or not description or not scope:
            raise EventLoadError(f"external_event_required_value:{sheet_name}:{offset}")
        rows.append(
            _native_event_row(
                snapshot_id,
                sheet_name,
                spec,
                offset,
                start,
                end,
                event_type,
                description,
                scope,
                raw,
            )
        )
    return rows


def _native_event_row(
    snapshot_id: str,
    sheet_name: str,
    spec: Mapping[str, Any],
    row_number: int,
    start: date,
    end: date,
    event_type: str,
    description: str,
    scope: str,
    raw: Sequence[Any],
) -> dict[str, Any]:
    identity = content_ref(
        "external-event",
        {"sheet": sheet_name, "values": [_json_cell(value) for value in raw]},
    )
    recurrence_spec = (
        spec.get("recurrence") if isinstance(spec.get("recurrence"), Mapping) else {}
    )
    recurrence = _parse_external_recurrence(
        description,
        recurrence_spec,
        sheet_name=sheet_name,
    )
    return _event_row(
        snapshot_id=snapshot_id,
        source_family="external_event",
        event_id=identity,
        event_type=str(event_type),
        start=start,
        end=end,
        scope=scope,
        authority="reviewed_workbook_pending_owner_review",
        evidence_level="context",
        wording_limit="context",
        payload={
            "sheet": sheet_name,
            "business_use": spec["business_use"],
            "description": description,
            "recurrence": dict(recurrence),
            "raw": [_json_cell(value) for value in raw],
        },
        recurrence=recurrence,
    )


def _event_row(
    *,
    snapshot_id: str,
    source_family: str,
    event_id: str,
    event_type: str,
    start: date,
    end: date,
    scope: str,
    authority: str,
    evidence_level: str,
    wording_limit: str,
    payload: Mapping[str, Any],
    recurrence: Mapping[str, Any] | None = None,
) -> dict[str, Any]:
    if not all((event_id, event_type, scope, authority, evidence_level, wording_limit)):
        raise EventLoadError("event_required_value")
    recurrence = recurrence or {}
    return {
        "snapshot_id": snapshot_id,
        "source_family": source_family,
        "event_id": event_id,
        "event_type": event_type,
        "event_start_date": start.isoformat(),
        "event_end_date": end.isoformat(),
        "affected_scope": scope,
        "authority": authority,
        "evidence_level": evidence_level,
        "wording_limit": wording_limit,
        "recurrence_kind": str(recurrence.get("kind") or ""),
        "recurrence_month_start": int(recurrence.get("month_start") or 0),
        "recurrence_day_start": int(recurrence.get("day_start") or 0),
        "recurrence_month_end": int(recurrence.get("month_end") or 0),
        "recurrence_day_end": int(recurrence.get("day_end") or 0),
        "payload": canonical_json_bytes(payload).decode("utf-8"),
    }


def _parse_external_recurrence(
    description: str,
    spec: Mapping[str, Any],
    *,
    sheet_name: str,
) -> dict[str, Any]:
    if not spec:
        return {}
    if spec.get("parser") != "reviewed_salary_rule_zh_v1":
        raise EventLoadError(f"external_event_recurrence_invalid:{sheet_name}")
    supported = set(spec.get("supported_kinds") or ())
    annual = re.search(
        r"(?P<start_month>\d{1,2})月\s*(?P<start_day>\d{1,2})日\s*[-~—至到]\s*"
        r"(?P<end_month>\d{1,2})月\s*(?P<end_day>\d{1,2})日",
        description,
    )
    if annual is not None:
        recurrence = {
            "kind": "annual_month_day_range",
            "month_start": int(annual["start_month"]),
            "day_start": int(annual["start_day"]),
            "month_end": int(annual["end_month"]),
            "day_end": int(annual["end_day"]),
            "timezone": str(spec.get("timezone") or ""),
        }
    else:
        monthly = re.search(
            r"(?:每月|次月)\s*(?P<start_day>\d{1,2})\s*[-~—至到]\s*"
            r"(?P<end_day>\d{1,2})日",
            description,
        )
        if monthly is None:
            raise EventLoadError(f"external_event_recurrence_unparsed:{sheet_name}")
        recurrence = {
            "kind": "monthly_day_range",
            "month_start": 0,
            "day_start": int(monthly["start_day"]),
            "month_end": 0,
            "day_end": int(monthly["end_day"]),
            "timezone": str(spec.get("timezone") or ""),
        }
    if recurrence["kind"] not in supported or not _valid_recurrence(recurrence):
        raise EventLoadError(f"external_event_recurrence_invalid:{sheet_name}")
    return recurrence


def _valid_recurrence(recurrence: Mapping[str, Any]) -> bool:
    kind = recurrence.get("kind")
    day_start = int(recurrence.get("day_start") or 0)
    day_end = int(recurrence.get("day_end") or 0)
    if kind == "monthly_day_range":
        return 1 <= day_start <= day_end <= 31
    if kind != "annual_month_day_range":
        return False
    try:
        date(2000, int(recurrence.get("month_start") or 0), day_start)
        date(2000, int(recurrence.get("month_end") or 0), day_end)
    except ValueError:
        return False
    return True


def _event_dates(value: Any, inferred_year: int | None) -> tuple[date, date, int]:
    if isinstance(value, datetime):
        return value.date(), value.date(), value.year
    if isinstance(value, date):
        return value, value, value.year
    if (
        isinstance(value, (int, float))
        and not isinstance(value, bool)
        and math.isfinite(float(value))
    ):
        from openpyxl.utils.datetime import from_excel

        parsed = from_excel(value)
        parsed_date = parsed.date() if isinstance(parsed, datetime) else parsed
        return parsed_date, parsed_date, parsed_date.year
    text = str(value or "").strip()
    quarter = re.search(r"(?P<year>20\d{2})\s*Q(?P<quarter>[1-4])", text, re.IGNORECASE)
    if quarter:
        year, number = int(quarter["year"]), int(quarter["quarter"])
        start = date(year, (number - 1) * 3 + 1, 1)
        end_month = number * 3
        next_month = date(
            year + (end_month == 12), 1 if end_month == 12 else end_month + 1, 1
        )
        return start, next_month - timedelta(days=1), year
    month_window = re.search(
        r"(?P<year>20\d{2})[-年/.](?P<start_month>\d{1,2})(?:月)?\s*(?:至|到|[-~—])\s*(?P<end_month>\d{1,2})月",
        text,
    )
    if month_window:
        year = int(month_window["year"])
        start_month = int(month_window["start_month"])
        end_month = int(month_window["end_month"])
        if not (
            1 <= start_month <= 12 and 1 <= end_month <= 12 and start_month <= end_month
        ):
            raise EventLoadError(f"external_event_date_invalid:{text}")
        return (
            date(year, start_month, 1),
            date(year, end_month, calendar.monthrange(year, end_month)[1]),
            year,
        )
    full = list(
        re.finditer(
            r"(?P<year>20\d{2})[-年/.](?P<month>\d{1,2})[-月/.](?P<day>\d{1,2})", text
        )
    )
    if full:
        start = date(int(full[0]["year"]), int(full[0]["month"]), int(full[0]["day"]))
        end = date(int(full[-1]["year"]), int(full[-1]["month"]), int(full[-1]["day"]))
        trailing = text[full[0].end() :]
        partial = list(
            re.finditer(
                r"(?<!\d)(?P<month>\d{1,2})[-月/.](?P<day>\d{1,2})(?!\d)", trailing
            )
        )
        if len(full) == 1 and partial:
            end = date(start.year, int(partial[-1]["month"]), int(partial[-1]["day"]))
            if end < start and end.month < start.month:
                end = date(start.year + 1, end.month, end.day)
        return start, end, start.year
    partial = list(
        re.finditer(r"(?<!\d)(?P<month>\d{1,2})[-月/.](?P<day>\d{1,2})(?!\d)", text)
    )
    if partial and inferred_year:
        start = date(inferred_year, int(partial[0]["month"]), int(partial[0]["day"]))
        end = date(inferred_year, int(partial[-1]["month"]), int(partial[-1]["day"]))
        return start, end, inferred_year
    raise EventLoadError(f"external_event_date_invalid:{text}")


def _strict_date(value: Any, field: str, error_type=EventLoadError) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    try:
        return date.fromisoformat(str(value).strip())
    except ValueError as exc:
        raise error_type(f"source_date_invalid:{field}") from exc


def _gameplay_filename(path: Path, *, channel: bool) -> tuple[date, date, str]:
    match = (GAMEPLAY_CHANNEL_RE if channel else GAMEPLAY_OVERALL_RE).fullmatch(
        path.name
    )
    if not match:
        raise GameplayLoadError(f"source_filename_invalid:{path.name}")
    start, end = date.fromisoformat(match["start"]), date.fromisoformat(match["end"])
    if start > end:
        raise GameplayLoadError(f"source_filename_window_invalid:{path.name}")
    channel_name = match["channel"].strip() if channel else ""
    if channel and not channel_name:
        raise GameplayLoadError(f"source_filename_invalid:{path.name}")
    return start, end, channel_name


def _validate_source_path(path: Path, seen_names: set[str], error_type) -> None:
    if not path.is_file():
        raise error_type(f"source_file_missing:{path}")
    if path.name in seen_names:
        raise error_type(f"duplicate_source_filename:{path.name}")
    seen_names.add(path.name)


def _gameplay_schema(
    specs: Mapping[str, Mapping[str, Any]], channel: bool
) -> tuple[str, ...]:
    fields = ["snapshot_id:String", "load_revision:String", "business_date:Date"]
    if channel:
        fields.append("channel:String")
    for field, spec in specs.items():
        if field == "business_date":
            continue
        data_type = str(spec["clickhouse_type"])
        if spec["nullable"]:
            data_type = f"Nullable({data_type})"
        fields.append(f"{field}:{data_type}")
    order = (
        "snapshot_id,load_revision,business_date,"
        + ("channel," if channel else "")
        + "service_scope,gameplay"
    )
    return (
        *fields,
        "engine:MergeTree",
        f"order_by:{order}",
        "canonicalization:gameplay-decimal-v1",
    )


def _event_schema(version: str) -> tuple[str, ...]:
    return (
        "snapshot_id:String",
        "load_revision:String",
        "source_family:LowCardinality(String)",
        "event_id:String",
        "event_type:String",
        "event_start_date:Date",
        "event_end_date:Date",
        "affected_scope:String",
        "authority:String",
        "evidence_level:String",
        "wording_limit:String",
        "recurrence_kind:String",
        "recurrence_month_start:UInt8",
        "recurrence_day_start:UInt8",
        "recurrence_month_end:UInt8",
        "recurrence_day_end:UInt8",
        "payload:String",
        "engine:MergeTree",
        "order_by:snapshot_id,load_revision,source_family,event_id",
        f"canonicalization:{version}",
    )


def _part(
    dataset_id: str,
    table_prefix: str,
    schema: tuple[str, ...],
    fingerprint: str,
    rows: Sequence[Mapping[str, Any]],
    date_range: tuple[str, ...],
    watermark: str,
    *,
    evidence_state: str,
    status: str = "active",
) -> DatasetLoadPart:
    return DatasetLoadPart(
        dataset_id,
        _versioned_table(table_prefix, fingerprint),
        tuple(
            item.split(":", 1)[0]
            for item in schema
            if not item.startswith(("engine:", "order_by:", "canonicalization:"))
        ),
        fingerprint,
        len(rows),
        tuple(date_range),
        watermark,
        rows_content_hash(rows),
        evidence_state,
        status,
        tuple(dict(row) for row in rows),
    )


def _build_manifest(
    *,
    snapshot_id: str,
    load_revision: str,
    source_family: str,
    contract_ref: str,
    canonicalization_version: str,
    source_checksums: Mapping[str, str],
    no_data_partitions: tuple[str, ...],
    no_data_partition_windows: tuple[str, ...],
    evidence_state: str,
    parts: tuple[DatasetLoadPart, ...],
) -> SourceLoadManifest:
    manifest_content = {
        "snapshot_id": snapshot_id,
        "load_revision": load_revision,
        "source_family": source_family,
        "contract_ref": contract_ref,
        "canonicalization_version": canonicalization_version,
        "source_checksums": source_checksums,
        "no_data_partitions": no_data_partitions,
        "no_data_partition_windows": no_data_partition_windows,
        "parts": [
            {key: value for key, value in asdict(part).items() if key != "rows"}
            for part in parts
        ],
    }
    manifest_ref = content_ref("source-load-manifest", manifest_content)
    snapshot_refs = tuple(
        content_ref(
            "dataset-snapshot",
            {
                "manifest_ref": manifest_ref,
                "snapshot_id": snapshot_id,
                "dataset_id": part.dataset_id,
                "physical_table": part.physical_table,
                "watermark": part.watermark,
                "schema_fingerprint": part.schema_fingerprint,
                "load_revision": load_revision,
            },
        )
        for part in parts
    )
    release_ref = dataset_snapshot_release_ref(
        snapshot_id, load_revision, snapshot_refs
    )
    return SourceLoadManifest(
        manifest_ref,
        snapshot_refs[0],
        snapshot_id,
        load_revision,
        release_ref,
        source_family,
        contract_ref,
        RUNTIME_BINDING_REF,
        canonicalization_version,
        dict(source_checksums),
        no_data_partitions,
        no_data_partition_windows,
        evidence_state,
        parts,
    )


def _snapshot_ref(manifest: SourceLoadManifest, part: DatasetLoadPart) -> str:
    return content_ref(
        "dataset-snapshot",
        {
            "manifest_ref": manifest.manifest_ref,
            "snapshot_id": manifest.snapshot_id,
            "dataset_id": part.dataset_id,
            "physical_table": part.physical_table,
            "watermark": part.watermark,
            "schema_fingerprint": part.schema_fingerprint,
            "load_revision": manifest.load_revision,
        },
    )


def _versioned_table(prefix: str, fingerprint: str) -> str:
    if not TABLE_RE.fullmatch(prefix) or not re.fullmatch(r"[0-9a-f]{64}", fingerprint):
        raise GameplayLoadError("invalid_versioned_table")
    return f"{prefix}__{fingerprint[:16]}"


def _attach_revision(
    rows: Sequence[Mapping[str, Any]], revision: str
) -> list[dict[str, Any]]:
    attached = []
    for item in rows:
        values = list(item.items())
        result = {values[0][0]: values[0][1], "load_revision": revision}
        result.update(dict(values[1:]))
        attached.append(result)
    return attached


def _row_date_range(rows: Sequence[Mapping[str, Any]], field: str) -> tuple[str, str]:
    values = tuple(str(item[field]) for item in rows)
    return min(values), max(values)


def _snapshot_available_at(watermark: str) -> str:
    return datetime.combine(
        date.fromisoformat(watermark) + timedelta(days=1), time.min, tzinfo=timezone.utc
    ).isoformat()


def _validate_event_identities(rows: Sequence[Mapping[str, Any]]) -> None:
    keys = [(str(row["source_family"]), str(row["event_id"])) for row in rows]
    if any(not key[1] for key in keys) or len(set(keys)) != len(keys):
        raise EventLoadError("event_identity_duplicate")


def _require_snapshot_id(snapshot_id: str, error_type) -> None:
    if not isinstance(snapshot_id, str) or not snapshot_id.strip():
        raise error_type("snapshot_id_required")


def _require_allowed(value: str, allowed: Sequence[str], code: str) -> None:
    if value not in allowed:
        raise EventLoadError(code)


def _json_cell(value: Any) -> Any:
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    if isinstance(value, Decimal):
        return format(value, "f")
    return value


def _schema_field_pair(value: str) -> tuple[str, str]:
    return tuple(value.split(":", 1))  # type: ignore[return-value]


def _read_persisted(
    client: Any,
    part: DatasetLoadPart,
    snapshot_id: str,
    load_revision: str,
) -> tuple[dict[str, Any], ...]:
    order = next(
        item.split(":", 1)[1]
        for item in _schema_for_part(part)
        if item.startswith("order_by:")
    )
    result = client.query(
        f"SELECT * FROM {part.physical_table} WHERE snapshot_id = {{snapshot_id:String}} AND load_revision = {{load_revision:String}} ORDER BY {order}",
        parameters={"snapshot_id": snapshot_id, "load_revision": load_revision},
        settings={"output_format_json_quote_decimals": 1},
    )
    raw_rows = tuple(dict(item) for item in result.named_results())
    expected_types = dict(
        _schema_field_pair(item)
        for item in _schema_for_part(part)
        if not item.startswith(("engine:", "order_by:", "canonicalization:"))
    )
    return tuple(
        _normalize_persisted_row(item, part, expected_types) for item in raw_rows
    )


def _schema_for_part(part: DatasetLoadPart) -> tuple[str, ...]:
    if part.dataset_id == GAMEPLAY_DATASET:
        return _gameplay_schema(
            _gameplay_field_specs(load_contract(GAMEPLAY_CONTRACT_PATH)), False
        )
    if part.dataset_id == GAMEPLAY_CHANNEL_DATASET:
        return _gameplay_schema(
            _gameplay_field_specs(load_contract(GAMEPLAY_CONTRACT_PATH)), True
        )
    return _event_schema(
        "external-events-v1"
        if part.dataset_id == EXTERNAL_EVENT_DATASET
        else "internal-operation-events-v1"
    )


def _normalize_persisted_row(
    row: Mapping[str, Any],
    part: DatasetLoadPart,
    expected_types: Mapping[str, str],
) -> dict[str, Any]:
    if set(row) != set(part.schema_fields):
        raise GameplayLoadError(f"persisted_columns_mismatch:{part.dataset_id}")
    normalized = {}
    for field in part.schema_fields:
        value = row[field]
        if value is None:
            normalized[field] = None
        elif expected_types[field] == "Date":
            normalized[field] = (
                value.isoformat() if isinstance(value, date) else str(value)
            )
        elif "Decimal" in expected_types[field]:
            scale_match = re.search(r"Decimal\(\d+,\s*(\d+)\)", expected_types[field])
            if scale_match is None:
                raise GameplayLoadError(f"persisted_decimal_type_invalid:{field}")
            normalized[field] = Decimal(str(value)).quantize(
                Decimal(1).scaleb(-int(scale_match.group(1))),
                rounding=ROUND_HALF_EVEN,
            )
        elif re.fullmatch(r"U?Int(?:8|16|32|64|128|256)", expected_types[field]):
            normalized[field] = int(value)
        else:
            normalized[field] = str(value)
    return normalized


def _validate_persisted_part(
    rows: Sequence[Mapping[str, Any]], part: DatasetLoadPart
) -> None:
    if len(rows) != part.row_count:
        raise GameplayLoadError(f"persisted_row_count_mismatch:{part.dataset_id}")
    schema = _schema_for_part(part)
    order = tuple(
        next(
            item.split(":", 1)[1] for item in schema if item.startswith("order_by:")
        ).split(",")
    )
    keys = [tuple(row[field] for field in order) for row in rows]
    if len(keys) != len(set(keys)):
        raise GameplayLoadError(f"persisted_unique_key_mismatch:{part.dataset_id}")
    if rows_content_hash(rows) != part.rows_content_hash:
        raise GameplayLoadError(f"persisted_rows_hash_mismatch:{part.dataset_id}")


def _gameplay_current_fingerprint(channel: bool) -> str:
    return schema_fingerprint(
        _gameplay_schema(
            _gameplay_field_specs(load_contract(GAMEPLAY_CONTRACT_PATH)), channel
        )
    )


def _event_current_fingerprint() -> str:
    return schema_fingerprint(_event_schema("external-events-v1"))


def main(argv: Sequence[str] | None = None) -> int:
    parser = argparse.ArgumentParser()
    subparsers = parser.add_subparsers(dest="command", required=True)
    gameplay_parser = subparsers.add_parser("gameplay")
    gameplay_parser.add_argument("--overall", type=Path, required=True)
    gameplay_parser.add_argument("--channels", type=Path, required=True)
    external_parser = subparsers.add_parser("external-events")
    external_parser.add_argument("--workbook", type=Path, required=True)
    internal_parser = subparsers.add_parser("internal-events")
    internal_parser.add_argument("--csv", type=Path, required=True)
    for item in (gameplay_parser, external_parser, internal_parser):
        item.add_argument("--snapshot-id", required=True)
        item.add_argument("--manifest-out", type=Path, required=True)
        item.add_argument("--clickhouse-container", default="")
        item.add_argument("--skip-postgres", action="store_true")
    args = parser.parse_args(argv)
    if args.command == "gameplay":
        if not args.overall.is_dir() or not args.channels.is_dir():
            raise GameplayLoadError("source_directory_missing")
        rows, manifest = load_gameplay_rows(
            tuple(args.overall.glob("*.csv")),
            tuple(args.channels.glob("*.csv")),
            snapshot_id=args.snapshot_id,
        )
    elif args.command == "external-events":
        rows, manifest = load_external_event_workbook(
            args.workbook, snapshot_id=args.snapshot_id
        )
    else:
        rows, manifest = load_internal_event_rows(
            args.csv, snapshot_id=args.snapshot_id
        )
    if args.clickhouse_container:
        database = os.environ.get("WAJE_CLICKHOUSE_DATABASE", "")
        if not database:
            raise GameplayLoadError(
                "missing_clickhouse_binding:WAJE_CLICKHOUSE_DATABASE"
            )
        client = DockerClickHouseClient(args.clickhouse_container, database)
    else:
        runtime = ClickHouseRuntime.from_env()
        if not runtime.configured():
            raise GameplayLoadError(
                "missing_clickhouse_binding:" + ",".join(runtime.binding.missing)
            )
        client = runtime._get_client()
    payloads = build_source_snapshot_payloads(manifest)
    store = None if args.skip_postgres else PostgresConversationStore.from_env()
    postgres_result = SnapshotPersistenceResult(active_refs=(), superseded_refs=())
    try:
        if store is None:
            apply_clickhouse_ddl(client, manifest)
            stage_source_release(client, manifest, active_load_revisions=())
        else:
            with store.dataset_snapshot_release_lock(manifest.snapshot_id):
                apply_clickhouse_ddl(client, manifest)
                active_revisions = tuple(
                    str(item.get("load_revision") or "")
                    for item in store.list_dataset_snapshots()
                    if item.get("logical_snapshot_id", item.get("snapshot_id"))
                    == manifest.snapshot_id
                    and item.get("status") == "active"
                )
                stage_source_release(
                    client, manifest, active_load_revisions=active_revisions
                )
                postgres_result = persist_source_snapshot_payloads(store, payloads)
    finally:
        if store is not None:
            close = getattr(store.connection, "close", None)
            if callable(close):
                close()
    artifact = {
        "source_load_manifest": manifest.to_dict(),
        "dataset_snapshot_payloads": list(
            postgres_result.verified_payloads or payloads
        ),
        "postgres_snapshot_roundtrip_refs": list(postgres_result.active_refs),
        "postgres_superseded_snapshot_refs": list(postgres_result.superseded_refs),
        "dataset_release_authority": postgres_result.authority_record,
    }
    args.manifest_out.parent.mkdir(parents=True, exist_ok=True)
    args.manifest_out.write_text(
        json.dumps(
            json.loads(canonical_json_bytes(artifact)),
            ensure_ascii=False,
            sort_keys=True,
            indent=2,
        )
        + "\n",
        encoding="utf-8",
    )
    print(
        json.dumps(
            {
                "manifest_path": str(args.manifest_out),
                "manifest_ref": manifest.manifest_ref,
                "row_counts": {
                    part.dataset_id: part.row_count for part in manifest.parts
                },
                "watermarks": {
                    part.dataset_id: part.watermark for part in manifest.parts
                },
                "release_ref": manifest.release_ref,
                "postgres_snapshot_refs": postgres_result.active_refs,
            },
            ensure_ascii=False,
            sort_keys=True,
        )
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
